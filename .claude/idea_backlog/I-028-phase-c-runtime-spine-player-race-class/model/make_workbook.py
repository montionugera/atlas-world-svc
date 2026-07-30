"""
Build combat-model.xlsx — every number is a LIVE FORMULA, not a pasted value.

Change anything yellow on the Inputs sheet (growth rate, stat coefficient, rank
multipliers, headcount) and the whole workbook recalculates: player curve, mob
stats, outcome matrix, and the pass/fail checks.

Run:  <venv>/bin/python make_workbook.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

wb = Workbook()

BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="2F4858")
INFILL = PatternFill("solid", fgColor="FFF2CC")      # yellow = editable input
CALCFILL = PatternFill("solid", fgColor="EAF1F8")    # blue = derived
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def header(ws, row, labels, widths=None):
    for i, t in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=t)
        c.font, c.fill, c.border = HDR, HDRFILL, THIN
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ───────────────────────────────────────────────────────────── Inputs
ws = wb.active
ws.title = "Inputs"
ws["A1"] = "COMBAT MODEL — INPUTS"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Yellow cells are editable. Everything else in the workbook recalculates from them."
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 62

INPUTS = [
    ("GROWTH", 1.045, "CombatScore growth per level (1.045 = +4.5%/lvl = 141% per 20)"),
    ("STATCOEF", 0.5, "C — a stat at its level cap grants +50%. Sets the stat share at C/(1+C)."),
    ("LVLMAX", 99, "Level cap"),
    ("BASEDPS", 20.0, "L1 anchor: DPS of a fully-allocated character"),
    ("BASEHP", 100.0, "L1 anchor: maxHP of a fully-allocated character"),
    ("BASEDEF", 5.0, "L1 anchor: defence (armour only — stats do NOT add to def)"),
    ("TARGETRED", 0.33, "Damage mitigation share of EHP. Held flat at every level by K(L)."),
    ("ATKSPD", 1.5, "Attacks per second"),
    ("CASTRATE", 0.8, "Casts per second"),
    ("PHYSMIX", 0.70, "Share of a reference build's DPS that is physical"),
    ("MSPDBASE", 20.0, "Move speed at zero AGI"),
    ("MSPDCAP", 36.0, "Hard clamp on move speed, applied AFTER all buffs"),
]
r = 4
ws.cell(row=r, column=1, value="constant").font = BOLD
ws.cell(row=r, column=2, value="value").font = BOLD
ws.cell(row=r, column=3, value="what it does").font = BOLD
for name, val, note in INPUTS:
    r += 1
    ws.cell(row=r, column=1, value=name).font = BOLD
    c = ws.cell(row=r, column=2, value=val)
    c.fill, c.border = INFILL, THIN
    ws.cell(row=r, column=3, value=note)
    wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$B${r}"))

r += 2
ws.cell(row=r, column=1, value="KANCHOR").font = BOLD
c = ws.cell(row=r, column=2, value="=BASEDEF*(1-TARGETRED)/TARGETRED")
c.fill, c.border = CALCFILL, THIN
ws.cell(row=r, column=3, value="K(1). K(L)=KANCHOR*growth — normalises defence so mitigation %% stays flat.")
wb.defined_names.add(DefinedName("KANCHOR", attr_text=f"Inputs!$B${r}"))

# player grades
r += 2
ws.cell(row=r, column=1, value="PLAYER GRADES").font = Font(bold=True, size=12)
r += 1
for i, t in enumerate(["grade", "allocation", "gear scale", ""], start=1):
    ws.cell(row=r, column=i, value=t).font = BOLD
GRADE_ROW = r + 1
for g, a, gs in [("max", 1.0, 1.00), ("median", 0.7, 0.85), ("min", 0.4, 0.70)]:
    r += 1
    ws.cell(row=r, column=1, value=g).font = BOLD
    for col, v in ((2, a), (3, gs)):
        c = ws.cell(row=r, column=col, value=v)
        c.fill, c.border = INFILL, THIN
ws.cell(row=r + 1, column=3,
        value="allocation = fraction of the level's stat cap actually invested; "
              "gear scale = fraction of best-in-slot")

# win bands
r += 3
ws.cell(row=r, column=1, value="WIN BANDS (R thresholds)").font = Font(bold=True, size=12)
r += 1
for i, t in enumerate(["band", "R below", "meaning"], start=1):
    ws.cell(row=r, column=i, value=t).font = BOLD
BAND_ROW = r + 1
for b, thr, mean in [("LOSS", 1.0, "you die first"),
                     ("brutal", 1.3, "<=23% HP left — a loss in practice"),
                     ("hard", 2.0, "23-50% HP left"),
                     ("fair", 3.5, "50-71% HP left"),
                     ("easy", 8.0, "71-87% HP left"),
                     ("trivial", 9999, ">=87% HP left")]:
    r += 1
    ws.cell(row=r, column=1, value=b).font = BOLD
    c = ws.cell(row=r, column=2, value=thr)
    c.fill, c.border = INFILL, THIN
    ws.cell(row=r, column=3, value=mean)
B_LOSS, B_BRUTAL, B_HARD, B_FAIR, B_EASY = (f"Inputs!$B${BAND_ROW+i}" for i in range(5))

ALLOC = {g: f"Inputs!$B${GRADE_ROW+i}" for i, g in enumerate(("max", "median", "min"))}
GEAR = {g: f"Inputs!$C${GRADE_ROW+i}" for i, g in enumerate(("max", "median", "min"))}


def verdict(cell):
    return (f'=IF({cell}<{B_LOSS},"LOSS",IF({cell}<{B_BRUTAL},"brutal",'
            f'IF({cell}<{B_HARD},"hard",IF({cell}<{B_FAIR},"fair",'
            f'IF({cell}<{B_EASY},"easy","trivial")))))')


# ───────────────────────────────────────────────────────────── Player
ws = wb.create_sheet("Player")
cols = ["L", "growth f", "K(L)", "DPS", "maxHP", "def", "mitigation", "EHP",
        "CS (max)", "pAtk", "mAtk", "mspd", "CS (median)", "CS (min)", "stat share"]
header(ws, 1, cols, [6, 11, 11, 11, 11, 10, 11, 12, 12, 10, 10, 9, 13, 12, 11])
for i in range(99):
    R = i + 2
    L = i + 1
    ws.cell(row=R, column=1, value=L)
    ws.cell(row=R, column=2, value=f"=GROWTH^(A{R}-1)")
    ws.cell(row=R, column=3, value=f"=KANCHOR*B{R}")
    # max grade
    ws.cell(row=R, column=4, value=f"=BASEDPS*B{R}/(1+STATCOEF)*(1+STATCOEF*{ALLOC['max']})*{GEAR['max']}")
    ws.cell(row=R, column=5, value=f"=BASEHP*B{R}/(1+STATCOEF)*(1+STATCOEF*{ALLOC['max']})*{GEAR['max']}")
    ws.cell(row=R, column=6, value=f"=BASEDEF*B{R}*{GEAR['max']}")
    ws.cell(row=R, column=7, value=f"=F{R}/(F{R}+C{R})")
    ws.cell(row=R, column=8, value=f"=E{R}/(1-G{R})")
    ws.cell(row=R, column=9, value=f"=SQRT(D{R}*H{R})")
    ws.cell(row=R, column=10, value=f"=D{R}*PHYSMIX/ATKSPD")
    ws.cell(row=R, column=11, value=f"=D{R}*(1-PHYSMIX)/CASTRATE")
    ws.cell(row=R, column=12, value=f"=MIN(MSPDBASE*(1+STATCOEF*{ALLOC['max']}),MSPDCAP)")
    for col, g in ((13, "median"), (14, "min")):
        d = f"BASEDPS*B{R}/(1+STATCOEF)*(1+STATCOEF*{ALLOC[g]})*{GEAR[g]}"
        h = f"BASEHP*B{R}/(1+STATCOEF)*(1+STATCOEF*{ALLOC[g]})*{GEAR[g]}"
        df = f"BASEDEF*B{R}*{GEAR[g]}"
        ws.cell(row=R, column=col, value=f"=SQRT(({d})*({h})/(1-({df})/(({df})+C{R})))")
    # stat share: 1 - CS(no stats) / CS(full), same gear
    d0 = f"BASEDPS*B{R}/(1+STATCOEF)"
    h0 = f"BASEHP*B{R}/(1+STATCOEF)"
    ws.cell(row=R, column=15,
            value=f"=1-SQRT(({d0})*({h0})/(1-F{R}/(F{R}+C{R})))/I{R}")
    ws.cell(row=R, column=7).number_format = "0%"
    ws.cell(row=R, column=15).number_format = "0.0%"
ws.freeze_panes = "A2"

# ───────────────────────────────────────────────────────────── Ladder
ws = wb.create_sheet("Ladder")
cols = ["rank", "lvl from", "lvl to", "ref lvl", "n (headcount)", "mult (per-mob)",
        "old mult", "ref CS", "mob CS", "mob HP", "mob pAtk", "mob pDef",
        "R solo (max)", "verdict", "R party (max)", "verdict",
        "R party (median)", "verdict", "R party (min)", "verdict",
        "TTK per mob", "TTK encounter"]
header(ws, 1, cols, [7, 9, 8, 8, 13, 14, 9, 10, 10, 11, 10, 10,
                     12, 10, 13, 10, 15, 10, 12, 10, 11, 12])
LADDER = [("E", 1, 12, 1, 0.29, 1.00), ("D", 13, 25, 1, 0.41, 1.15),
          ("C", 26, 40, 1, 0.50, 1.30), ("B", 41, 55, 2, 0.78, 1.50),
          ("A", 56, 70, 4, 0.943, 1.80), ("S", 71, 84, 8, 1.054, 2.20),
          ("SS", 85, 95, 20, 1.127, 2.80), ("SSS", 96, 99, 50, 1.183, 3.50)]
for i, (rank, l0, l1, n, mult, old) in enumerate(LADDER):
    R = i + 2
    ws.cell(row=R, column=1, value=rank).font = BOLD
    ws.cell(row=R, column=2, value=l0)
    ws.cell(row=R, column=3, value=l1)
    ws.cell(row=R, column=4, value=f"=ROUNDDOWN((B{R}+C{R})/2,0)")
    for col, v in ((5, n), (6, mult)):
        c = ws.cell(row=R, column=col, value=v)
        c.fill, c.border = INFILL, THIN
    ws.cell(row=R, column=7, value=old)
    ws.cell(row=R, column=8, value=f"=INDEX(Player!$I:$I,D{R}+1)")
    ws.cell(row=R, column=9, value=f"=H{R}*F{R}")
    ws.cell(row=R, column=10, value=f"=INDEX(Player!$E:$E,D{R}+1)*F{R}")
    ws.cell(row=R, column=11, value=f"=INDEX(Player!$J:$J,D{R}+1)*F{R}")
    ws.cell(row=R, column=12, value=f"=INDEX(Player!$F:$F,D{R}+1)*F{R}")
    # R_solo = (CS_p/CS_m)^2 ; R_party = R_solo * 2n/(n+1)
    ws.cell(row=R, column=13, value=f"=(H{R}/I{R})^2")
    ws.cell(row=R, column=14, value=verdict(f"M{R}"))
    ws.cell(row=R, column=15, value=f"=M{R}*2*E{R}/(E{R}+1)")
    ws.cell(row=R, column=16, value=verdict(f"O{R}"))
    ws.cell(row=R, column=17,
            value=f"=(INDEX(Player!$M:$M,D{R}+1)/I{R})^2*2*E{R}/(E{R}+1)")
    ws.cell(row=R, column=18, value=verdict(f"Q{R}"))
    ws.cell(row=R, column=19,
            value=f"=(INDEX(Player!$N:$N,D{R}+1)/I{R})^2*2*E{R}/(E{R}+1)")
    ws.cell(row=R, column=20, value=verdict(f"S{R}"))
    # TTK: n players focus one mob, then the whole pack
    ws.cell(row=R, column=21,
            value=f"=(INDEX(Player!$H:$H,D{R}+1)*F{R})/(INDEX(Player!$D:$D,D{R}+1)*E{R})")
    ws.cell(row=R, column=22, value=f"=U{R}*E{R}")
    for col in (13, 15, 17, 19, 21, 22):
        ws.cell(row=R, column=col).number_format = "0.00"
ws.freeze_panes = "B2"

# ───────────────────────────────────────────────────────────── Checks
ws = wb.create_sheet("Checks")
ws["A1"] = "REQUIREMENTS & INVARIANTS"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "All of these recalculate when you change Inputs or the Ladder multipliers."
header(ws, 4, ["#", "check", "target", "actual", "result"], [4, 58, 26, 14, 10])

# ladder row lookup by rank
def lrow(rank):
    return 2 + [r[0] for r in LADDER].index(rank)


CHECKS = [
    ("max player CANNOT solo a same-level S mob", "R < 1.0",
     f"=Ladder!M{lrow('S')}", f"=IF(Ladder!M{lrow('S')}<1,\"PASS\",\"FAIL\")"),
    ("max player cannot EASILY solo same-level A", f"R < fair",
     f"=Ladder!M{lrow('A')}", f"=IF(Ladder!M{lrow('A')}<{B_FAIR},\"PASS\",\"FAIL\")"),
    ("median player beats same-level C, fair", "hard <= R < fair",
     f"=Ladder!Q{lrow('C')}",
     f"=IF(AND(Ladder!Q{lrow('C')}>={B_HARD},Ladder!Q{lrow('C')}<{B_FAIR}),\"PASS\",\"FAIL\")"),
    ("max player beats same-level C, easy", "fair <= R < easy",
     f"=Ladder!O{lrow('C')}",
     f"=IF(AND(Ladder!O{lrow('C')}>={B_FAIR},Ladder!O{lrow('C')}<{B_EASY}),\"PASS\",\"FAIL\")"),
    ("20-level growth inside 120-160%", "1.20 - 1.60",
     "=GROWTH^20-1", "=IF(AND(GROWTH^20-1>=1.2,GROWTH^20-1<=1.6),\"PASS\",\"FAIL\")"),
    ("CS compounds at the growth rate", "= GROWTH",
     "=Player!I3/Player!I2", "=IF(ABS(Player!I3/Player!I2-GROWTH)<0.000001,\"PASS\",\"FAIL\")"),
    ("mitigation % flat across all levels", "max - min = 0",
     "=MAX(Player!G2:G100)-MIN(Player!G2:G100)",
     "=IF(MAX(Player!G2:G100)-MIN(Player!G2:G100)<0.000001,\"PASS\",\"FAIL\")"),
    ("stat share >= 25% at EVERY level", ">= 25%",
     "=MIN(Player!O2:O100)", "=IF(MIN(Player!O2:O100)>=0.25,\"PASS\",\"FAIL\")"),
    ("stat share does not decay with level", "max - min = 0",
     "=MAX(Player!O2:O100)-MIN(Player!O2:O100)",
     "=IF(MAX(Player!O2:O100)-MIN(Player!O2:O100)<0.000001,\"PASS\",\"FAIL\")"),
    ("every rank is winnable by its own party size", "all R party (max) > 1",
     "=MIN(Ladder!O2:O9)", "=IF(MIN(Ladder!O2:O9)>1,\"PASS\",\"FAIL\")"),
    ("S and above are NOT soloable", "all R solo < 1 from S up",
     f"=MAX(Ladder!M{lrow('S')}:M{lrow('SSS')})",
     f"=IF(MAX(Ladder!M{lrow('S')}:M{lrow('SSS')})<1,\"PASS\",\"FAIL\")"),
    ("largest mob HP fits int32", "< 2,147,483,647",
     "=MAX(Ladder!J2:J9)", "=IF(MAX(Ladder!J2:J9)<2147483647,\"PASS\",\"FAIL\")"),
    ("move speed within its clamp", "<= MSPDCAP",
     "=MAX(Player!L2:L100)", "=IF(MAX(Player!L2:L100)<=MSPDCAP,\"PASS\",\"FAIL\")"),
]
for i, (name, target, actual, res) in enumerate(CHECKS):
    R = 5 + i
    ws.cell(row=R, column=1, value=i + 1)
    ws.cell(row=R, column=2, value=name)
    ws.cell(row=R, column=3, value=target)
    a = ws.cell(row=R, column=4, value=actual)
    a.number_format = "0.000"
    c = ws.cell(row=R, column=5, value=res)
    c.font, c.alignment = BOLD, Alignment(horizontal="center")
R = 5 + len(CHECKS) + 1
ws.cell(row=R, column=2, value="OVERALL").font = Font(bold=True, size=12)
ws.cell(row=R, column=5,
        value=f'=IF(COUNTIF(E5:E{4+len(CHECKS)},"FAIL")=0,"ALL PASS",'
              f'COUNTIF(E5:E{4+len(CHECKS)},"FAIL")&" FAILING")').font = Font(bold=True, size=12)

# ───────────────────────────────────────────────────────────── Readme
ws = wb.create_sheet("Readme", 0)
ws.column_dimensions["A"].width = 110
lines = [
    ("Atlas combat model — live workbook", 14, True),
    ("", 11, False),
    ("Every figure is a formula. Change the yellow cells and the whole book moves.", 11, False),
    ("", 11, False),
    ("Inputs   the constants. GROWTH and STATCOEF are the two that reshape everything.", 11, False),
    ("Player   the L1-99 curve: DPS, HP, defence, mitigation, EHP, CS at three player grades.", 11, False),
    ("Ladder   per-rank mob strength and the outcome R for solo and for a correctly-sized party.", 11, False),
    ("Checks   your four stated requirements plus the model invariants, recomputed live.", 11, False),
    ("", 11, False),
    ("The core model", 12, True),
    ("  pAtk  = (base + weapon) x (1 + C x str/statCap(L))     stats MULTIPLY gear, never add", 11, False),
    ("  maxHP = (base + armour) x (1 + C x vit/statCap(L))", 11, False),
    ("  pDef / mDef = armour only, so mitigation %% is allocation-independent", 11, False),
    ("  CS    = SQRT(DPS x EHP)", 11, False),
    ("", 11, False),
    ("The outcome metric", 12, True),
    ("  R = how long you survive / how long the encounter survives.  R>1 = win.", 11, False),
    ("  HP left at victory = 1 - 1/R", 11, False),
    ("  R_solo  = (CS_player / CS_mob)^2        <- CS is the exact sufficient statistic", 11, False),
    ("  R_party = R_solo x 2n/(n+1)             <- n players vs n mobs, Lanchester both ways", 11, False),
    ("", 11, False),
    ("Known gaps — deliberately NOT in this workbook", 12, True),
    ("  Mana, skills and physical-vs-magic parity (see mana_level.py, parity.py).", 11, False),
    ("  Crit, misses, kiting, AoE. Closed-form only; a BattleModule run is still owed.", 11, False),
    ("  Jobs: R cancels the DPS/EHP split, so allocation cannot change any verdict here.", 11, False),
    ("  TTK at high ranks derives to seconds, not the 3000s+ the old table assumed —", 11, False),
    ("  open question whether top ranks should be n players vs ONE boss instead.", 11, False),
]
for i, (t, sz, bold) in enumerate(lines, start=1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = Font(bold=bold, size=sz)

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "combat-model.xlsx")
wb.save(out)
print(f"wrote {out}")
